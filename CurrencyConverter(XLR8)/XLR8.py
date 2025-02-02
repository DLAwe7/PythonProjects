import CalculoEconomico as cl
import openpyxl as xl

wb = xl.load_workbook("Calculo economico.xlsx")
sheet = wb["Feuille1"]


print("Welcome to the currencies converter.")

input_currency = cl.get_currency_input("Starting currency: ")
input_amount = cl.get_currency_amount("Quantity: ")
output_currency = cl.get_currency_input("Desired currency: ")


rate_from = None
rate_to = None


for col_idx, col in enumerate(sheet.iter_cols(min_col=1, max_col=sheet.max_column), start=1):
    for row_idx, cell in enumerate(col, start=1):
        if cell.value == input_currency:
            cell = sheet.cell(row=row_idx, column=col_idx - 1)
            rate_from = float(cell.value)
            break


for col_idx, col in enumerate(sheet.iter_cols(min_col=1, max_col=sheet.max_column), start=1):
    for row_idx, cell in enumerate(col, start=1):
        if cell.value == output_currency:
            cell = sheet.cell(row=row_idx, column=col_idx - 1)
            rate_to = float(cell.value)
            break

total = input_amount / rate_to * rate_from
print(f"{input_amount} {input_currency} is equivalent to {total:.2f} {output_currency}.")







