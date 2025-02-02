import openpyxl as xl

wb = xl.load_workbook("Calculo economico.xlsx")
sheet = wb["Feuille1"]


def get_currency_input(prompt):
    while True:
        original_currency = input(prompt)
        input_currency = original_currency.upper()
        if input_currency not in [cell.value for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column) for cell in col]:
            print(f"'{original_currency}' is not a valid input. Please insert a currency/ISO 4217 code (EUR, USD...).")
        else:
            break
    return input_currency


def get_currency_amount(prompt):
    while True:
        try:
            temp_amount = input(prompt)
            input_amount = float(temp_amount)
            break
        except ValueError:
            print(f"{temp_amount} is not a valid number. Please try again.")
    return input_amount


def add_currency(code, rate_to_eur):

    code = code.upper()
    if not isinstance(rate_to_eur, (int, float)):
        print(f"'{rate_to_eur}' is not a valid number. Please try again")
        return
    if len(code) != 3:
        print(f"'{code}' is not a valid input. Please insert a currency/ISO 4217 code (EUR, USD...).")
        return

    new_row_idx = sheet.max_row + 1
    sheet.cell(row=new_row_idx, column=4).value = code
    sheet.cell(row=new_row_idx, column=3).value = rate_to_eur
    print(f"New currency successfully added on row {new_row_idx} in columns 4 and 3")
    wb.save("Calculo economico.xlsx")


def del_last_currency():
    last_row = None
    for row_idx in range(sheet.max_row, 0, -1):
        if sheet.cell(row=row_idx, column=4).value is not None:
            last_row = row_idx
            break

    if last_row is not None:

        sheet.cell(row=last_row, column=4).value = None
        sheet.cell(row=last_row, column=3).value = None
        print(f"Deleted currency from row {last_row}.")
    else:
        print("No currency to delete.")


    wb.save("Calculo economico.xlsx")
    print("Workbook saved after deletion.")


def del_specific_currency(code):
    code = code.upper()

    if len(code) != 3:
        print(f"'{code}' is not a valid input. Please insert a currency/ISO 4217 code (EUR, USD...).")
        return


    found = False

    for col_idx, col in enumerate(sheet.iter_cols(min_col=1, max_col=sheet.max_column), start=1):
        for row_idx, cell in enumerate(col, start=-1):
            if cell.value == code:
                sheet.cell(row=row_idx, column=col_idx).value = None
                sheet.cell(row=row_idx, column=3).value = None
                print(f"Deleted {code} from excel sheet.")
                found = True
                wb.save("Calculo economico.xlsx")
                return


    if not found:
        print(f"'{code}' was not found within {sheet}.")
        return


